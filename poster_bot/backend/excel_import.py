"""
Excel Import - парсинг групп из Excel файла
"""
import re
from typing import List, Tuple
from openpyxl import load_workbook
from io import BytesIO


def extract_invite_hash(link: str) -> Tuple[str, str]:
    """
    Извлекает тип и хеш/username из ссылки на группу

    Returns:
        (link_type, value) - 'invite' + hash или 'public' + username
    """
    link = link.strip()

    # Приватная ссылка: t.me/+xxxxx или t.me/joinchat/xxxxx
    invite_match = re.search(r't\.me/\+([a-zA-Z0-9_-]+)', link)
    if invite_match:
        return ('invite', invite_match.group(1))

    joinchat_match = re.search(r't\.me/joinchat/([a-zA-Z0-9_-]+)', link)
    if joinchat_match:
        return ('invite', joinchat_match.group(1))

    # Публичная ссылка: t.me/channelname
    public_match = re.search(r't\.me/([a-zA-Z][a-zA-Z0-9_]{3,})', link)
    if public_match:
        username = public_match.group(1)
        # Исключаем служебные пути
        if username.lower() not in ['joinchat', 'addstickers', 'share']:
            return ('public', username)

    return (None, None)


def normalize_link(link: str) -> str:
    """Нормализует ссылку к единому формату"""
    link = link.strip()

    # Убираем http/https
    link = re.sub(r'^https?://', '', link)

    # Добавляем https://
    if not link.startswith('http'):
        link = 'https://' + link

    return link


def detect_columns(sheet) -> dict:
    """
    Определяет какие колонки содержат нужные данные

    Returns:
        dict с ключами: city_col, address_col, link_col, has_header
    """
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    second_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)

    result = {
        'city_col': None,
        'address_col': None,
        'link_col': None,
        'has_header': False
    }

    if not first_row:
        return result

    # Ищем по заголовкам
    for idx, cell_value in enumerate(first_row):
        if not cell_value or not isinstance(cell_value, str):
            continue

        header = cell_value.lower().strip()

        # Город
        if header in ['город', 'city', 'населенный пункт']:
            result['city_col'] = idx
            result['has_header'] = True

        # Адрес
        elif header in ['адрес', 'address', 'улица']:
            result['address_col'] = idx
            result['has_header'] = True

        # Ссылка
        elif header in ['ссылка', 'link', 'url', 'ссылка на чат', 'чат', 'группа', 'channel']:
            result['link_col'] = idx
            result['has_header'] = True

    # Если нашли заголовки, проверяем первую колонку без заголовка
    if result['has_header'] and second_row:
        # Если первая колонка пустая в заголовке но есть данные - это город
        if first_row[0] is None and second_row[0] and result['city_col'] is None:
            # Проверяем что это не ссылка
            if not (isinstance(second_row[0], str) and 't.me/' in second_row[0]):
                result['city_col'] = 0

    # Если заголовки не найдены, пробуем определить по данным
    if result['link_col'] is None:
        # Ищем колонку со ссылками
        for idx, cell_value in enumerate(first_row):
            if cell_value and isinstance(cell_value, str):
                if 't.me/' in cell_value.lower():
                    result['link_col'] = idx
                    break

        # Если нашли ссылку в первой строке - это не заголовок
        if result['link_col'] is not None:
            result['has_header'] = False
            # Пытаемся определить город и адрес по позиции
            # Обычно: город - 0, адрес - 1, ссылка - 2
            if result['link_col'] >= 2:
                result['city_col'] = 0
                result['address_col'] = 1
            elif result['link_col'] == 1:
                result['address_col'] = 0

    return result


def parse_excel(file_content: bytes) -> List[dict]:
    """
    Парсит Excel файл и извлекает ссылки на группы с дополнительной информацией

    Returns:
        List[dict] - список групп с полями: link, city, address, link_type, invite_hash
    """
    workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
    sheet = workbook.active

    groups = []
    seen_links = set()

    # Определяем структуру колонок
    columns = detect_columns(sheet)

    link_col = columns['link_col']
    city_col = columns['city_col']
    address_col = columns['address_col']
    has_header = columns['has_header']

    if link_col is None:
        # Не нашли колонку со ссылками
        workbook.close()
        return groups

    # С какой строки начинать
    start_row = 2 if has_header else 1

    # Парсим строки
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if not row or len(row) <= link_col:
            continue

        # Ссылка
        link_value = row[link_col]
        if not link_value:
            continue

        link = str(link_value).strip()
        if not link or 't.me' not in link.lower():
            continue

        # Нормализуем ссылку
        link = normalize_link(link)

        # Проверяем дубликаты
        if link in seen_links:
            continue
        seen_links.add(link)

        # Извлекаем тип и хеш
        link_type, value = extract_invite_hash(link)
        if not link_type:
            continue

        # Город
        city = None
        if city_col is not None and len(row) > city_col and row[city_col]:
            city = str(row[city_col]).strip()

        # Адрес
        address = None
        if address_col is not None and len(row) > address_col and row[address_col]:
            address = str(row[address_col]).strip()

        groups.append({
            'link': link,
            'city': city,
            'address': address,
            'link_type': link_type,
            'invite_hash': value if link_type == 'invite' else None,
            'username': value if link_type == 'public' else None
        })

    workbook.close()
    return groups


def validate_excel(file_content: bytes) -> Tuple[bool, str, int]:
    """
    Валидирует Excel файл

    Returns:
        (is_valid, error_message, count)
    """
    try:
        groups = parse_excel(file_content)
        if not groups:
            return (False, "Не найдено ни одной ссылки на группу/канал", 0)
        return (True, None, len(groups))
    except Exception as e:
        return (False, f"Ошибка чтения файла: {str(e)}", 0)
