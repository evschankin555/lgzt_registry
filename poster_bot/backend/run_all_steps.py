"""
Полный цикл: retry FloodWait -> join -> send unsent -> retry failed sends -> final report (Excel)
"""
import requests
import time
import sys
import io
from datetime import datetime, timezone, timedelta

# Fix Windows console encoding for emoji/unicode
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

API_BASE = "https://lgzt.developing-site.ru/api"
PASSWORD = "admin2026" + chr(36) + "rassilka"
PHONE = "+79936932613"
MESSAGE_ID = 1

MSK = timezone(timedelta(hours=3))


def auth():
    r = requests.post(f"{API_BASE}/login", json={"password": PASSWORD})
    r.raise_for_status()
    token = r.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def get_stats(h):
    stats = requests.get(f"{API_BASE}/groups/stats", headers=h).json()
    print(f"  joined={stats['joined']}, left={stats['left']}, "
          f"failed={stats['failed']}, pending={stats['pending']}, total={stats['total']}")
    return stats


# ============================================================
# AUTH
# ============================================================
print("=" * 60)
print("Авторизация...")
h = auth()
print("OK\n")

# ============================================================
# ШАГ 1: Сброс FloodWait групп в pending
# ============================================================
print("=" * 60)
print("ШАГ 1: Сброс FloodWait групп в pending")
print("=" * 60)

r = requests.post(f"{API_BASE}/groups/reset-failed",
                  data={"error_contains": "wait of"}, headers=h)
result = r.json()
print(f"  Сброшено: {result['reset_count']} из {result['total_failed']} failed")

# Также сбросим "requested to join" — они тоже могли быть одобрены
r2 = requests.post(f"{API_BASE}/groups/reset-failed",
                   data={"error_contains": "requested to join"}, headers=h)
result2 = r2.json()
print(f"  Сброшено 'requested to join': {result2['reset_count']}")

get_stats(h)
print()

# ============================================================
# ШАГ 2: Запуск join worker
# ============================================================
print("=" * 60)
print("ШАГ 2: Запуск join worker для pending групп")
print("=" * 60)

stats = requests.get(f"{API_BASE}/groups/stats", headers=h).json()
pending_count = stats["pending"]

if pending_count == 0:
    print("  Нет pending групп, пропускаем join")
else:
    print(f"  Pending групп: {pending_count}, запускаем join worker...")
    r = requests.post(f"{API_BASE}/groups/start-joining",
                      json={"phone": PHONE, "limit": 0, "delay_min": 30, "delay_max": 60},
                      headers=h)
    print(f"  Join worker: {r.json()}")

    # Ждём завершения
    wait_iter = 0
    while True:
        time.sleep(15)
        wait_iter += 1

        try:
            status = requests.get(f"{API_BASE}/groups/joining-status", headers=h).json()
            is_running = status.get("is_running", False)
            jstats = status.get("stats", {})

            joined_session = jstats.get("joined_this_session", 0)
            pending_left = jstats.get("pending", 0)
            failed_session = jstats.get("failed_this_session", 0)
            next_in = jstats.get("next_attempt_in", 0)

            sys.stdout.write(
                f"\r  [{wait_iter}] joined={joined_session}, "
                f"failed={failed_session}, pending={pending_left}, "
                f"next={next_in}s, running={is_running}    "
            )
            sys.stdout.flush()

            if not is_running:
                print(f"\n  Join worker завершён!")
                break

            # Таймаут 60 минут
            if wait_iter > 240:
                print(f"\n  Таймаут ожидания, останавливаем...")
                requests.post(f"{API_BASE}/groups/stop-joining", headers=h)
                time.sleep(5)
                break

        except Exception as e:
            print(f"\n  Ошибка проверки статуса: {e}")
            time.sleep(10)

    print()
    get_stats(h)

print()

# ============================================================
# ШАГ 3: Отправка сообщений в joined группы без отправки
# ============================================================
print("=" * 60)
print("ШАГ 3: Отправка в joined группы без отправки")
print("=" * 60)

# Получаем все группы и message targets
all_groups = requests.get(f"{API_BASE}/groups", headers=h).json()
msg_detail = requests.get(f"{API_BASE}/messages/{MESSAGE_ID}", headers=h).json()

sent_group_ids = set()
for s in msg_detail.get("sends", []):
    sent_group_ids.add(s["group_id"])

# Группы joined без отправки
unsent = [g for g in all_groups
          if g["status"] == "joined"
          and g["id"] not in sent_group_ids
          and g.get("telegram_id")]

print(f"  Найдено {len(unsent)} joined групп без отправки")

if unsent:
    # Отправляем по 1 группе, задержка 120 сек между отправками на клиенте
    total_ok = 0
    total_fail = 0

    for idx, g in enumerate(unsent):
        # Задержка 120 сек между отправками (кроме первой)
        if idx > 0:
            print(f"    Ждём 120 сек...")
            time.sleep(120)

        title = g.get("title") or g.get("address") or g["link"]
        print(f"\n  [{idx+1}/{len(unsent)}] {title}")

        try:
            resp = requests.post(
                f"{API_BASE}/messages/{MESSAGE_ID}/send",
                data={"phone": PHONE, "group_ids": str(g["id"]), "delay_seconds": 0},
                headers=h, timeout=120
            )
            result = resp.json()

            for r in result.get("results", []):
                if r["status"] == "sent":
                    print(f"    OK")
                    total_ok += 1
                else:
                    err = (r.get("error") or "")[:60]
                    print(f"    FAIL: {err}")
                    total_fail += 1
        except Exception as e:
            print(f"    ERROR: {e}")
            total_fail += 1

    print(f"\n  Итого новых: OK={total_ok}, FAIL={total_fail}")
else:
    print("  Нечего отправлять")

print()

# ============================================================
# ШАГ 4: Повторная отправка в failed группы
# ============================================================
print("=" * 60)
print("ШАГ 4: Повторная отправка в failed группы")
print("=" * 60)

# Перечитываем данные
msg_detail = requests.get(f"{API_BASE}/messages/{MESSAGE_ID}", headers=h).json()
all_groups = requests.get(f"{API_BASE}/groups", headers=h).json()
groups_map = {g["id"]: g for g in all_groups}

failed_sends = [s for s in msg_detail.get("sends", []) if s["status"] == "failed"]
print(f"  Найдено {len(failed_sends)} failed отправок")

retry_groups = []
for s in failed_sends:
    g = groups_map.get(s["group_id"])
    if g and g.get("telegram_id"):
        retry_groups.append(g)

print(f"  С telegram_id: {len(retry_groups)}")

if retry_groups:
    total_ok = 0
    total_fail = 0
    still_failed = []

    for idx, g in enumerate(retry_groups):
        # Задержка 120 сек между отправками (кроме первой)
        if idx > 0:
            print(f"    Ждём 120 сек...")
            time.sleep(120)

        title = g.get("title") or g.get("address") or g["link"]
        print(f"\n  [{idx+1}/{len(retry_groups)}] {title}")

        try:
            resp = requests.post(
                f"{API_BASE}/messages/{MESSAGE_ID}/send",
                data={"phone": PHONE, "group_ids": str(g["id"]), "delay_seconds": 0},
                headers=h, timeout=120
            )
            result = resp.json()

            for r in result.get("results", []):
                if r["status"] == "sent":
                    print(f"    OK")
                    total_ok += 1
                else:
                    err = (r.get("error") or "")[:60]
                    print(f"    FAIL: {err}")
                    total_fail += 1
                    still_failed.append({"title": title, "error": r.get("error", "")})
        except Exception as e:
            print(f"    ERROR: {e}")
            total_fail += 1

    print(f"\n  Итого retry: OK={total_ok}, FAIL={total_fail}")
    if still_failed:
        print(f"  Безнадёжные ({len(still_failed)}):")
        for sf in still_failed:
            print(f"    - {sf['title']}: {sf['error'][:60]}")
else:
    print("  Нечего повторять")

print()

# ============================================================
# ШАГ 5: Финальный отчёт (Excel)
# ============================================================
print("=" * 60)
print("ШАГ 5: Формирование финального отчёта")
print("=" * 60)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Перечитываем финальные данные
h = auth()  # refresh token
stats = requests.get(f"{API_BASE}/groups/stats", headers=h).json()
all_groups = requests.get(f"{API_BASE}/groups", headers=h).json()
msg_detail = requests.get(f"{API_BASE}/messages/{MESSAGE_ID}", headers=h).json()
groups_map = {g["id"]: g for g in all_groups}

print(f"  Финальная статистика:")
get_stats(h)

# Категоризация отправок
successful_sends = []
failed_sends_list = []

for s in msg_detail.get("sends", []):
    group = groups_map.get(s["group_id"], {})
    entry = {
        "group_title": s.get("group_title", "N/A"),
        "group_link": s.get("group_link") or group.get("link", "N/A"),
        "city": group.get("city", ""),
        "address": group.get("address", ""),
        "message_link": s.get("message_link", ""),
        "sent_at": s.get("sent_at", ""),
        "error": s.get("error", ""),
        "status": s.get("status", "")
    }
    if s["status"] == "sent":
        successful_sends.append(entry)
    elif s["status"] == "failed":
        failed_sends_list.append(entry)

# Категоризация групп
failed_join = [g for g in all_groups if g["status"] == "failed"]
pending_groups = [g for g in all_groups if g["status"] == "pending"]
left_groups = [g for g in all_groups if g["status"] == "left"]
joined_groups = [g for g in all_groups if g["status"] == "joined"]

expired_groups = [g for g in failed_join if "expired" in (g.get("join_error") or "")]
flood_groups = [g for g in failed_join if "wait of" in (g.get("join_error") or "").lower()]
request_groups = [g for g in failed_join if "requested to join" in (g.get("join_error") or "").lower()]
invalid_link = [g for g in failed_join if any(x in (g.get("join_error") or "").lower() for x in ["no user has", "cannot cast"])]
other_failed = [g for g in failed_join
                if g not in expired_groups and g not in flood_groups
                and g not in request_groups and g not in invalid_link]

# Стили
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def fmt_date(iso_str):
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d.%m.%Y %H:%M")
    except:
        return str(iso_str)


wb = openpyxl.Workbook()

# ---- Лист 1: Сводка ----
ws = wb.active
ws.title = "Сводка"
ws.append(["Итоговый отчёт по рассылке"])
ws.cell(1, 1).font = Font(bold=True, size=14)
now = datetime.now(MSK)
ws.append([f"Дата отчёта: {now.strftime('%d.%m.%Y %H:%M МСК')}"])
ws.append([])
ws.append(["Показатель", "Значение"])
style_header(ws, 4, 2)

rows = [
    ("Всего групп в базе", stats["total"]),
    ("Успешно отправлено", len(successful_sends)),
    ("Не удалось отправить (нет прав)", len(failed_sends_list)),
    ("Сейчас в группах (joined)", stats["joined"]),
    ("Вышли из групп", stats["left"]),
    ("Не вошли — истёкшие ссылки", len(expired_groups)),
    ("Не вошли — FloodWait", len(flood_groups)),
    ("Не вошли — запрос одобрения", len(request_groups)),
    ("Не вошли — невалидные ссылки", len(invalid_link)),
    ("Не вошли — другие ошибки", len(other_failed)),
    ("В очереди (pending)", stats["pending"]),
]
for label, val in rows:
    ws.append([label, val])

# Подсветка
for r in range(5, ws.max_row + 1):
    ws.cell(r, 1).border = thin_border
    ws.cell(r, 2).border = thin_border
    if ws.cell(r, 1).value and "Успешно" in str(ws.cell(r, 1).value):
        ws.cell(r, 2).fill = green_fill

auto_width(ws)

# ---- Лист 2: Все ОК (успешно вступили + отправили) ----
ws2 = wb.create_sheet("Всё ОК")
ws2.append(["N", "Группа", "Город", "Адрес", "Ссылка на группу", "Ссылка на сообщение", "Дата отправки"])
style_header(ws2, 1, 7)

# Все группы, где отправка прошла успешно (sent) — и вышли, и ещё в группе
for i, s in enumerate(successful_sends, 1):
    ws2.append([i, s["group_title"], s["city"], s["address"],
                s["group_link"], s["message_link"], fmt_date(s["sent_at"])])

auto_width(ws2)

# ---- Лист 3: Успешно отправлено (ещё в группе) ----
ws3 = wb.create_sheet("Успешно отправлено")
ws3.append(["N", "Группа", "Город", "Адрес", "Ссылка на группу", "Ссылка на сообщение", "Дата отправки"])
style_header(ws3, 1, 7)

# Только те, кто ещё joined или left + sent
for i, s in enumerate(successful_sends, 1):
    ws3.append([i, s["group_title"], s["city"], s["address"],
                s["group_link"], s["message_link"], fmt_date(s["sent_at"])])

auto_width(ws3)

# ---- Лист 4: Не удалось отправить ----
ws4 = wb.create_sheet("Не удалось отправить")
ws4.append(["N", "Группа", "Город", "Ссылка на группу", "Ошибка"])
style_header(ws4, 1, 5)

for i, s in enumerate(failed_sends_list, 1):
    ws4.append([i, s["group_title"], s["city"], s["group_link"], s["error"][:120]])

auto_width(ws4)

# ---- Лист 5: Невалидные ссылки ----
ws5 = wb.create_sheet("Невалидные ссылки")
ws5.append(["N", "Ссылка", "Город", "Адрес"])
style_header(ws5, 1, 4)

for i, g in enumerate(expired_groups + invalid_link, 1):
    ws5.append([i, g["link"], g.get("city", ""), g.get("address", "")])

auto_width(ws5)

# ---- Лист 6: FloodWait ----
ws6 = wb.create_sheet("FloodWait")
ws6.append(["N", "Ссылка", "Город", "Адрес", "Ошибка"])
style_header(ws6, 1, 5)

for i, g in enumerate(flood_groups, 1):
    ws6.append([i, g["link"], g.get("city", ""), g.get("address", ""), (g.get("join_error") or "")[:80]])

auto_width(ws6)

# ---- Лист 7: Запрос одобрения ----
ws7 = wb.create_sheet("Запрос одобрения")
ws7.append(["N", "Ссылка", "Город", "Адрес", "Ошибка"])
style_header(ws7, 1, 5)

for i, g in enumerate(request_groups, 1):
    ws7.append([i, g["link"], g.get("city", ""), g.get("address", ""), (g.get("join_error") or "")[:80]])

auto_width(ws7)

# ---- Лист 8: В ожидании ----
ws8 = wb.create_sheet("В ожидании")
ws8.append(["N", "Ссылка", "Город", "Адрес"])
style_header(ws8, 1, 4)

for i, g in enumerate(pending_groups, 1):
    ws8.append([i, g["link"], g.get("city", ""), g.get("address", "")])

auto_width(ws8)

# ---- Лист 9: Вышли из групп ----
ws9 = wb.create_sheet("Вышли из групп")
ws9.append(["N", "Группа", "Ссылка", "Город", "Дата выхода"])
style_header(ws9, 1, 5)

for i, g in enumerate(left_groups, 1):
    ws9.append([i, g.get("title", "N/A"), g["link"], g.get("city", ""), fmt_date(g.get("left_at"))])

auto_width(ws9)

# Сохраняем
filename = f"poster_report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(filename)

print(f"\n  Отчёт сохранён: {filename}")
print(f"  Листы: {wb.sheetnames}")
print(f"\n  Успешно отправлено: {len(successful_sends)}")
print(f"  Не удалось: {len(failed_sends_list)}")
print(f"  Невалидные: {len(expired_groups) + len(invalid_link)}")
print(f"  FloodWait: {len(flood_groups)}")
print(f"  Запрос одобрения: {len(request_groups)}")
print(f"  Pending: {len(pending_groups)}")
print(f"  Вышли: {len(left_groups)}")

print("\n" + "=" * 60)
print("ГОТОВО!")
print("=" * 60)
